import os
import glob
import zipfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, numbers


class CreateExcel:
    def __init__(self, task):
        self.date = task.date
        self.date_year = self.date.strftime('Y')
        self.date_month = self.date.strftime('m')
        self.date_day = self.date.strftime('d')
        self.random_str = task.file_no
        self.date_str = self.date.strftime('%Y-%m-%d')
        self.filepath = self.date.strftime('media{0}baobiao{0}%Y-%m{0}%d{0}'.format(os.sep))
        
        self.filename_content = '{}-{}'.format(self.date_str, self.random_str)
        self.filename = '{}-汇总.xlsx'.format(self.filename_content)
        self.all_center = Alignment(horizontal='center', vertical='center')
        self.workbook = Workbook()

        # 创建一张新表
        self.worksheet = self.workbook.active
        self.header()
        try:
            os.makedirs(self.filepath)
            # print()
        except Exception as e:
            pass

    def close(self):
        self.workbook.save("{}{}".format(self.filepath, self.filename))

    def header(self):
        """
        """
        # 填充汇总表 文字为居中
        self.worksheet['A1'].alignment = self.all_center
        self.worksheet['B1'].alignment = self.all_center
        self.worksheet['C1'].alignment = self.all_center
        # 添加表头
        self.worksheet['A1'] = '表格名称'
        self.worksheet['B1'] = '汇总金额(万元)'
        self.worksheet['C1'] = '笔数'
        # 调整表格宽度
        self.worksheet.column_dimensions['A'].width = 25+2
        self.worksheet.column_dimensions['B'].width = 12+2
        self.worksheet.column_dimensions['C'].width = 4+2
    
    def insert(self, instance):
        i = str(instance.num+1)
        # 填充汇总表数据
        self.worksheet['A%s'% i ] = '{}-{}.xlsx'.format(self.filename_content, instance.num)
        self.worksheet['B%s'% i ] = instance.amount_total
        self.worksheet['C%s'% i ] = instance.batch_total

        # 填充汇总表 文字为居中
        self.worksheet['A%s'% i ].alignment = self.all_center
        self.worksheet['B%s'% i ].alignment = self.all_center
        self.worksheet['C%s'% i ].alignment = self.all_center

    def footer(self):
        pass

    def creat_zipfile(self):
        file_list = glob.iglob(r'{}*{}*.xlsx'.format(self.filepath, self.random_str))
        z = zipfile.ZipFile('{}{}.zip'.format(self.filepath, self.filename_content), 'w') 
        for file in file_list:
            z.write(file, file.replace(self.filepath, ''))
        z.close()


class TranExcel(CreateExcel):
    def __init__(self, taskbatch):
        self.task = taskbatch.task
        self.taskbatch = taskbatch
        super(TranExcel,self).__init__(self.task)
        self.filename_content = '{}-{}-转账文件-{}'.format(self.date_str, self.random_str, taskbatch.num)
        self.filename = '{}.xlsx'.format(self.filename_content)

    def close(self):
        self.workbook.save("{}{}".format(self.filepath, self.filename))

    def header(self):
        pass

    def footer(self):
        pass


class FujianTranExcel(TranExcel):
    """
    """
    def header(self):
        """
        """
        super(FujianTranExcel,self).header()
        # 添加表头
        self.worksheet['A1'] = '账户账号'
        self.worksheet['B1'] = '转账金额'
        self.worksheet['C1'] = '备注'
        # 调整表格宽度
        self.worksheet.column_dimensions['A'].width = 15+2
        self.worksheet.column_dimensions['B'].width = 12+2
        self.worksheet.column_dimensions['C'].width = 10+2
    
    def insert(self, i, instance, account):
        """
        """
        self.worksheet['A%s'% i] = account.account
        self.worksheet['A%s'% i].alignment = self.all_center
        self.worksheet['A%s'% i].number_format = numbers.FORMAT_TEXT
        self.worksheet['B%s'% i] = instance.amount * 10000
        self.worksheet['B%s'% i].alignment = self.all_center
        self.worksheet['C%s'% i] = '商户资金结算: %s' % instance.order_no
        self.worksheet['C%s'% i].alignment = self.all_center


class JiangxiTranExcel(TranExcel):
    """
    """
    def header(self):
        """
        """
        self.worksheet.merge_cells('A1:G1')
        self.worksheet.cell(row=1, column=1).value = '付款单\r\n         {}'.format(self.task.date.strftime('%Y年%m月%d日'))
        self.worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        self.worksheet.row_dimensions[1].height = 40

        self.worksheet['A2'] = '序号'
        self.worksheet['B2'] = '收款方户名'
        self.worksheet['C2'] = '收款方账户号'
        self.worksheet['D2'] = '收款账户开户行名称'
        self.worksheet['E2'] = '收款账号开户行行号'
        self.worksheet['F2'] = '付款金额'
        self.worksheet['G2'] = '资金用途'

        # 设置居中效果
        self.worksheet['A2'].alignment = self.all_center
        self.worksheet['B2'].alignment = self.all_center
        self.worksheet['C2'].alignment = self.all_center
        self.worksheet['D2'].alignment = self.all_center
        self.worksheet['E2'].alignment = self.all_center
        self.worksheet['F2'].alignment = self.all_center
        self.worksheet['G2'].alignment = self.all_center
        
        # 调整表格宽度
        self.worksheet.column_dimensions['A'].width = 6*2+2
        self.worksheet.column_dimensions['B'].width = 15+2
        self.worksheet.column_dimensions['C'].width = 15+2
        self.worksheet.column_dimensions['D'].width = 15+2
        self.worksheet.column_dimensions['E'].width = 15+2
        self.worksheet.column_dimensions['F'].width = 12+2
        self.worksheet.column_dimensions['G'].width = 10+2

    def insert(self, j, instance, account):
        """
        """
        self.worksheet['A%s'% str(j+1)] = j-1
        self.worksheet['A%s'% str(j+1)].alignment = self.all_center

        self.worksheet['B%s'% str(j+1)] = account.account_name
        self.worksheet['B%s'% str(j+1)].alignment = self.all_center

        self.worksheet['C%s'% str(j+1)] = account.account
        self.worksheet['C%s'% str(j+1)].alignment = self.all_center

        self.worksheet['D%s'% str(j+1)] = account.bank_name
        self.worksheet['D%s'% str(j+1)].alignment = self.all_center

        self.worksheet['E%s'% str(j+1)] = account.bank_code
        self.worksheet['E%s'% str(j+1)].alignment = self.all_center

        self.worksheet['F%s'% str(j+1)] = instance.amount * 10000
        self.worksheet['F%s'% str(j+1)].alignment = self.all_center

        self.worksheet['G%s'% str(j+1)] = '商户资金结算: %s' % instance.order_no
        self.worksheet['G%s'% str(j+1)].alignment = self.all_center

        # 设置单元格格式为文本
        self.worksheet['B%s'% str(j+1)].number_format = numbers.FORMAT_TEXT
        self.worksheet['E%s'% str(j+1)].number_format = numbers.FORMAT_TEXT

    def footer(self):
        j = self.taskbatch.batch_total + 3
        self.worksheet['A%s'% str(j)] = '合计（笔数）'
        self.worksheet['A%s'% str(j)].alignment = self.all_center
        self.worksheet['B%s'% str(j)] = j - 3 
        self.worksheet['B%s'% str(j)].alignment = self.all_center

        self.worksheet['E%s'% str(j)] = '合计（金额）'
        self.worksheet['E%s'% str(j)].alignment = self.all_center
        self.worksheet['F%s'% str(j)] = self.taskbatch.amount_total * 10000
        self.worksheet['F%s'% str(j)].alignment = self.all_center
    
    def close(self):
        self.footer()
        self.workbook.save("{}{}".format(self.filepath, self.filename))

class TranInfoExcel(CreateExcel):
    def __init__(self, task):
        self.task = task
        super(TranInfoExcel,self).__init__(self.task)
        self.filename_content = '{}-{}-转账信息'.format(self.date_str, self.random_str)
        self.filename = '{}.xlsx'.format(self.filename_content)

    def close(self):
        self.workbook.save("{}{}".format(self.filepath, self.filename))
        pass

    def header(self):
        """
        """
        # 填充汇总表 文字为居中
        self.worksheet['A1'].alignment = self.all_center
        self.worksheet['B1'].alignment = self.all_center
        self.worksheet['C1'].alignment = self.all_center
        self.worksheet['D1'].alignment = self.all_center
        self.worksheet['E1'].alignment = self.all_center
        self.worksheet['F1'].alignment = self.all_center
        self.worksheet['G1'].alignment = self.all_center
        self.worksheet['H1'].alignment = self.all_center
        self.worksheet['I1'].alignment = self.all_center
        self.worksheet['J1'].alignment = self.all_center
        self.worksheet['K1'].alignment = self.all_center
        self.worksheet['L1'].alignment = self.all_center
        self.worksheet['M1'].alignment = self.all_center
        self.worksheet['N1'].alignment = self.all_center

        self.worksheet['O1'].alignment = self.all_center
        self.worksheet['P1'].alignment = self.all_center
        self.worksheet['Q1'].alignment = self.all_center
        self.worksheet['R1'].alignment = self.all_center
        self.worksheet['S1'].alignment = self.all_center

        self.worksheet['T1'].alignment = self.all_center
        self.worksheet['U1'].alignment = self.all_center
        self.worksheet['V1'].alignment = self.all_center
        self.worksheet['W1'].alignment = self.all_center
        self.worksheet['x1'].alignment = self.all_center

        
        # 添加表头
        self.worksheet['A1'] = '编号'
        self.worksheet['B1'] = '产品ID'
        self.worksheet['C1'] = '商家名称'
        self.worksheet['D1'] = '商家ID'
        self.worksheet['E1'] = '子公司名称'
        self.worksheet['F1'] = '子公司ID'
        self.worksheet['G1'] = '母公司名称'
        self.worksheet['H1'] = '母公司ID'
        self.worksheet['I1'] = '买家名称'
        self.worksheet['J1'] = '买家ID'
        self.worksheet['K1'] = '订单号'
        self.worksheet['L1'] = '商品状态'
        self.worksheet['M1'] = '下单时商品名称'
        self.worksheet['N1'] = '下单时商品规格'
        self.worksheet['O1'] = '下单时价格'
        self.worksheet['P1'] = '订单金额(总额元)'
        self.worksheet['Q1'] = '订金金额(万元)'
        self.worksheet['R1'] = '下单时商品类型'
        self.worksheet['S1'] = '下单数量'
        self.worksheet['T1'] = '下单时单位'
        self.worksheet['U1'] = '订单金额比例'
        self.worksheet['V1'] = '创建时间'
        self.worksheet['W1'] = '修改时间'
        self.worksheet['X1'] = '删除标记'



        # 调整表格宽度
        self.worksheet.column_dimensions['A'].width = 25+2
        self.worksheet.column_dimensions['B'].width = 20+2
        self.worksheet.column_dimensions['C'].width = 25+2
        self.worksheet.column_dimensions['D'].width = 12+2
        self.worksheet.column_dimensions['E'].width = 12+2
        self.worksheet.column_dimensions['F'].width = 25+2
        self.worksheet.column_dimensions['G'].width = 5+2
        self.worksheet.column_dimensions['H'].width = 5+2
        self.worksheet.column_dimensions['I'].width = 25+2
        self.worksheet.column_dimensions['J'].width = 25+2
        self.worksheet.column_dimensions['K'].width = 25+2
        self.worksheet.column_dimensions['L'].width = 15+2
        self.worksheet.column_dimensions['M'].width = 5+2
        self.worksheet.column_dimensions['N'].width = 10+2
        self.worksheet.column_dimensions['O'].width = 5+2
        self.worksheet.column_dimensions['P'].width = 5+2
        self.worksheet.column_dimensions['Q'].width = 25+2
        self.worksheet.column_dimensions['R'].width = 25+2
        self.worksheet.column_dimensions['S'].width = 25+2
        self.worksheet.column_dimensions['T'].width = 15+2
        self.worksheet.column_dimensions['U'].width = 5+2
        self.worksheet.column_dimensions['V'].width = 10+2
        self.worksheet.column_dimensions['W'].width = 15+2
        self.worksheet.column_dimensions['X'].width = 5+2


    def insert(self, i, instance):
        """
        """
        self.worksheet['C%s'% i] = instance.seller.name
        self.worksheet['E%s'% i] = instance.buyer.company.name
        self.worksheet['I%s'% i] = instance.buyer.name
        self.worksheet['K%s'% i] = instance.order_no
        self.worksheet['M%s'% i] = instance.products.name
        self.worksheet['N%s'% i] = instance.products.type 
        self.worksheet['O%s'% i].number_format = numbers.FORMAT_TEXT
        self.worksheet['O%s'% i] = instance.price
        self.worksheet['P%s'% i].number_format = numbers.FORMAT_TEXT
        self.worksheet['P%s'% i] = instance.tran_tatal
        self.worksheet['Q%s'% i].number_format = numbers.FORMAT_TEXT
        self.worksheet['Q%s'% i] = instance.amount
        self.worksheet['R%s'% i] =  instance.products.scope.name
        self.worksheet['S%s'% i] = instance.quantity
        self.worksheet['S%s'% i].number_format = numbers.FORMAT_TEXT
        self.worksheet['T%s'% i] = instance.products.unit
        self.worksheet['V%s'% i] = instance.date.strftime('%Y/%m/%d')


        self.worksheet['C%s'% i].alignment = self.all_center
        self.worksheet['E%s'% i].alignment = self.all_center
        self.worksheet['I%s'% i].alignment = self.all_center
        self.worksheet['K%s'% i].alignment = self.all_center
        self.worksheet['M%s'% i].alignment = self.all_center
        self.worksheet['N%s'% i].alignment = self.all_center
        self.worksheet['O%s'% i].alignment = self.all_center
        self.worksheet['P%s'% i].alignment = self.all_center
        self.worksheet['Q%s'% i].alignment = self.all_center
        self.worksheet['R%s'% i].alignment = self.all_center
        self.worksheet['S%s'% i].alignment = self.all_center
        self.worksheet['T%s'% i].alignment = self.all_center
        self.worksheet['V%s'% i].alignment = self.all_center


def test():
    from Tran.models import Task, TaskBatch, Transaction
    from Account.models import Account
    task = Task.objects.get(id=2)
    tran_huizong_excel = CreateExcel(task)
    for i, taskbatch in enumerate(TaskBatch.objects.filter(task=task), 1):
        tran_huizong_excel.insert(taskbatch)
        tran_excel = FujianTranExcel(taskbatch)
        traninfo_excel = TranInfoExcel(taskbatch)
        # print(tran_excel.filename)
        for j, transaction in enumerate(Transaction.objects.filter(task_batch=taskbatch), 2):
            # 转账记录表&转账信息表插入一条数据
            account = Account.objects.filter(company=transaction.buyer.company).order_by('?').first()
            tran_excel.insert(j, transaction, account)
            traninfo_excel.insert(j, transaction)
        tran_excel.close()
        traninfo_excel.close()
    tran_huizong_excel.close()
    tran_huizong_excel.creat_zipfile()


if __name__ == '__main__':
    test()
    